#! /usr/bin/env python
from openpyxl import load_workbook
from ics import Calendar, Event
from datetime import date, time, datetime, timedelta
from pytz import timezone

FILE_NAME = "edt.xlsx"

BEGIN_OFFSET = 4
END_OFFSET = 238
WEEK_ROW_SIZE = 9
HALF_DAY_ROW_SIZE = 4

BEGIN_WEEK = 2
END_WEEK = 11
DAY_SIZE = 2

DAYS = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"]

wb = load_workbook(FILE_NAME)
ws = wb[wb.get_sheet_names()[0]]

PARIS_TZ = timezone("Europe/Paris")
SCHOOL_DAY = date(2016, 9, 12)

cal = Calendar()

timetable = {
        "FF66FFFF": {"morning": [time(8, 45), time(12, 30)],
            "afternoon": [time(13, 45), time(17, 30)]},
        # Seems to be in error in a cell
        "FF00FFFF": {"morning": [time(8, 45), time(12, 30)],
            "afternoon": [time(13, 45), time(17, 30)]},
        "FFFFFF00": {"morning": [time(9, 0), time(12, 0)],
            "afternoon": [time(13, 30), time(16, 30)]},
        "FFFF33FF": {"morning": [time(9, 0), time(12, 30)],
            "afternoon": [time(13, 30), time(17, 0)]},
        "FF00FF00": {"morning": [time(9, 0), time(12, 30)],
            "afternoon": [time(13, 30), time(17, 0)]},
        "FF4A86E8": {"morning": [time(8, 45), time(12, 30)],
            "afternoon": [time(13, 45), time(17, 30)]},
        "FFF9CB9C": {"morning": [time(8, 45), time(12, 30)],
            "afternoon": [time(13, 45), time(17, 30)]},
        "FFFF0066": {"morning": [time(9, 0), time(12, 30)],
            "afternoon": [time(14, 0), time(17, 0)]},
        "default": {"morning": [time(8, 45), time(12, 30)],
            "afternoon": [time(13, 45), time(17, 30)]},
        }

def add_cell_as_event_to_call(cal, day, cell, half_day):
    if cell.value is None:
        return
    if cell.fill.bgColor.rgb in timetable:
        cal.events.append(Event(name=cell.value,
            begin=PARIS_TZ.localize(datetime.combine(day, timetable[cell.fill.bgColor.rgb][half_day][0])),
            end=PARIS_TZ.localize(datetime.combine(day, timetable[cell.fill.bgColor.rgb][half_day][1])),
            ))
    else:
        print("No hours found for")
        print(cell.value)
        print("Color = " + cell.fill.bgColor.rgb)
        cal.events.append(Event(name=cell.value,
            begin=PARIS_TZ.localize(datetime.combine(day, timetable["default"][half_day][0])),
            end=PARIS_TZ.localize(datetime.combine(day, timetable["default"][half_day][1])),
            ))

current_day = SCHOOL_DAY
for j in range(BEGIN_OFFSET, END_OFFSET, WEEK_ROW_SIZE):
    for i in range(BEGIN_WEEK, END_WEEK + 1, DAY_SIZE):
        # Morning classes
        cell = ws.cell(row = j, column = i)
        add_cell_as_event_to_call(cal, current_day, cell, "morning")
        cell = ws.cell(row = j, column = i + 1)
        add_cell_as_event_to_call(cal, current_day, cell, "morning")

        # Afternoon classes
        cell = ws.cell(row = j + HALF_DAY_ROW_SIZE, column = i)
        add_cell_as_event_to_call(cal, current_day, cell, "afternoon")
        cell = ws.cell(row = j + HALF_DAY_ROW_SIZE, column = i + 1)
        add_cell_as_event_to_call(cal, current_day, cell, "afternoon")

        current_day = current_day + timedelta(1)
    current_day = current_day + timedelta(2)

print("\nSaving")
with open("cal.ics", "w") as f:
    f.writelines(cal)
print("Done!")
